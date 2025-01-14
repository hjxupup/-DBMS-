import hashlib
import os
import re
import time
from openpyxl import *
from prettytable import PrettyTable
import dbms_function
from dbms_function import (
    create_tb_in_tbinfo,
    check_permission,
    insert_record,
)

db_path = 'data/'
# view_path = 'view/'
user = ''
using_dbname = ''
using_db = Workbook()

def help():
    """
    打印帮助信息
    :return:
    """
    print("""
    右下角输入指令，submit等于回车

    点击数据库按钮刷新指令

    点击全选选择所有数据库

    点击加载加载数据

    加载完成后可在data中查看每个数据库以及sheet

    ## 登录管理员
    username:admin
    username:admin

    ## 创建数据库
    create database {database_name}
    eg.: create database test_db

    ## 删除数据库
    drop database {database_name}
    eg.: drop database test_db

    ## 使用数据库
    use database {database_name}
    eg.: use database test_db

    ## 创建表
    create table {table_name} ({column_name} {data_type} {PK,null...},{column_name} {data_type} {PK,null...}...)
    eg.: create table test (v1 int PK null,v2 int)

    ## 删除表
    drop table {table_name}
    eg.: drop table test

    ## 添加字段
    alter {table_name} add ({column_name} {data_type} {PK,null...})
    eg.: alter test add (v3 int)

    ## 删除字段
    alter {table_name} drop ({column_name})
    eg.: alter test drop (v3)

    ## 修改字段
    alter {table_name} modify {alter_field_name} ({column_name} {data_type} {PK,null...}) 
    eg.: alter test modify v1 (v3 int PK null)
    
    ## 记录插入
    insert into {table_name} {column_name=value,column_name=value,...)
    eg.: insert into test v1=1,v2=2

    ## 记录插入（多重）
    insert into {table_name} {column_name=value,column_name=value,...&column_name=value,column_name=value,...)
    eg.: insert into test v3=2,v2=4&v3=3,v2=5

    ## 记录删除
    delete on {table_name} where {column_name=value或column_name>value或column_name<value}
    eg.: delete on test where v3=1

    ## 记录删除（多重）
    delete on {table_name} where {column_name=value或column_name>value或column_name<value&column_name=value或column_name>value或column_name<value&..}
    eg.: delete on test where v3=1&v2=2

    ## 记录修改
    update {table_name} set column_name=value,column_name=value,... where {column_name=value或column_name>value或column_name<value（可多重）}
    eg.: update test set v3=4,v2=3 where v3=2

    ## 选择全部
    select * from {table_name}
    eg.: select * from test

    ## 选择指定列
    select {column_name} from {table_name}
    eg.:select v3 from test

    ## 选择where条件
    select * 或{column_name} from {table_name} where {column_name=value或column_name>value或column_name<value（可多重）}
    eg.: select * from test where v3=4

    ## 注册用户
    signup {username} {password}
    eg.: signup admin admin

    ## 读取脚本
    load {script_name}
    eg.: load test.txt

    ## 创建视图
    create view {view_name} as select * 或{column_name} from {table_name}
    eg.: create view test as select * from test

    ## 赋予权限
    grant {action} on {database_name} for {username}
    eg.: grant select on test_db for aaa

    ## 收回权限
    revoke {action} on {database_name} for {username}
    eg.: revoke select on test_db for aaa

    """)


def use_db(dbname):
    global using_dbname
    global using_db
    if os.path.exists(db_path + dbname + '.xlsx'):
        if dbms_function.check_permission(user, dbname, 'use'):
            using_dbname = dbname
            print(dbname + "数据库已使用.")
            using_db = load_workbook(db_path + dbname + '.xlsx')
        else:
            print("你没有权限使用该数据库,请使用admin账户赋予权限.")
    else:
        print("数据库不存在")

def show_db():
    print("All database:")
    dbs = os.listdir(db_path)
    for db in dbs:
        if '.DS' not in db and db != 'index':
            print("[*] " + db[:-5])

def creat_db(dbname):
    dbpath = 'data/' + dbname + '.xlsx'
    database = Workbook()
    database.save(dbpath)
    dbms_function.create_tb_in_tbinfo(dbname)
    print(u"数据库创建成功")

def get_command():
    command = input("[👉]> ") if not using_dbname else input("[{}🚩]> ".format(using_dbname))
    return command.strip()

def Initialization():
    if not os.path.exists(db_path):
        os.mkdir(db_path)
    if not os.path.exists("data/table_information.xlsx"):
        Workbook().save("data/table_information.xlsx")
    if os.path.exists("data/system.xlsx"):
        print("Initializating......")
    else:
        dbms_function.creat_db('system')
    db = load_workbook("data/system.xlsx")
    permission_tb_col = ['database char[50] pk unique', 'select char', 'insert char', 'delete char', 'update char']
    dbms_function.creat_table('permission', db, 'system', permission_tb_col)

def query(sql, tag=''):
    print(f"解析SQL: {sql}")
    sql_word = sql.split(" ")
    global using_dbname
    global using_db
    
    if len(sql_word) < 2:
        print("[!] Wrong query!")
        return
        
    operate = sql_word[0].lower()
    print(f"操作类型: {operate}")
    
    if operate == 'use':
        if sql_word[1] == 'database':
            try:
                use_db(sql_word[2])
            except:
                print("[!]Error")
        else:
            print("[!]Syntax Error.\neg:>use database dbname")
            
    elif operate == 'create':
        if sql_word[1] == 'database':
            try:
                creat_db(sql_word[2])
            except:
                print("[!]Create Error")
        elif sql_word[1] == 'table':
            columns_list = re.findall(r'\((.*?)\)', sql)[0].split(',')
            print(columns_list, using_dbname)
            try:
                dbms_function.creat_table(sql_word[2], using_db, using_dbname, columns_list)
                using_db = load_workbook(db_path + using_dbname + '.xlsx')
            except:
                print("[!]Error")
                
    elif operate == 'load':
        if len(sql_word) < 2:
            print("[!] 请指定脚本文件名")
            return
            
        script_name = sql_word[1]
        script_path = os.path.join(os.getcwd(), 'data/script', script_name)
        
        if not os.path.exists(script_path):
            print(f"[!] 脚本文件 {script_name} 不存在")
            return
            
        import sys
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        print(f"[{timestamp}] [调试] 脚本路径: {script_path}", file=sys.stderr)
        print(f"[{timestamp}] [调试] 脚本路径: {script_path}", file=sys.stdout)
        try:
            with open(script_path, 'r', encoding='utf-8') as f:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                print(f"[{timestamp}] [调试] 成功打开脚本文件", file=sys.stderr)
                print(f"[{timestamp}] [调试] 成功打开脚本文件", file=sys.stdout)
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#'):  # 忽略空行和注释
                        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        print(f"[{timestamp}] [执行] {line}", file=sys.stderr)
                        print(f"[{timestamp}] [执行] {line}", file=sys.stdout)
                        query(line)
                    else:
                        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        print(f"[{timestamp}] [调试] 跳过行: {line}", file=sys.stderr)
                        print(f"[{timestamp}] [调试] 跳过行: {line}", file=sys.stdout)
        except Exception as e:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            print(f"[{timestamp}] [!] 读取脚本文件时出错: {str(e)}", file=sys.stderr)
            print(f"[{timestamp}] [!] 读取脚本文件时出错: {str(e)}", file=sys.stdout)
            print(f"[{timestamp}] [调试] 当前工作目录: {os.getcwd()}", file=sys.stderr)
            print(f"[{timestamp}] [调试] 当前工作目录: {os.getcwd()}", file=sys.stdout)
            
    elif operate == 'insert':
        print("处理插入操作")
        # 解析表名和值
        match = re.match(r'insert\s+into\s+(\w+)\s+(.*)', sql)
        if match:
            table_name = match.group(1)
            values_str = match.group(2)
            print(f"表名: {table_name}")
            print(f"值: {values_str}")
            
            # 解析键值对
            pairs = values_str.split(',')
            columns_list = []
            for pair in pairs:
                if '=' in pair:
                    key, value = pair.split('=')
                    columns_list.append([key.strip(), value.strip()])
                    
            print(f"解析后的列表: {columns_list}")
            
            # 检查是否有权限
            if not check_permission(user, using_dbname, 'insert'):
                print("你没有插入权限")
                return
                
            # 执行插入
            insert_record(table_name, using_db, using_dbname, columns_list, False)
            # 保存更改
            using_db.save(db_path + using_dbname + '.xlsx')
            print("插入成功")
        else:
            print("[!] Invalid INSERT syntax")

    # 处理SELECT查询
    if operate == 'select':
        try:
            # 检查是否有查询权限
            if not check_permission(user, using_dbname, 'select'):
                print("你没有查询权限")
                return
                
            # 解析JOIN查询
            if 'join' in sql.lower():
                # 解析表名和连接条件
                tables = []
                join_conditions = []
                
                # 提取第一个表名
                from_idx = sql_word.index('from')
                tables.append(sql_word[from_idx + 1])
                
                # 提取其他表名和连接条件
                join_idx = sql_word.index('join')
                tables.append(sql_word[join_idx + 1])
                
                # 提取连接条件
                on_idx = sql_word.index('on')
                join_condition = sql_word[on_idx + 1]
                join_conditions.append(join_condition)
                
                # 提取 WHERE 条件
                where_condition = None
                if 'where' in sql_word:
                    where_idx = sql_word.index('where')
                    where_condition = ' '.join(sql_word[where_idx + 1:])
                
                # 提取要查询的列
                columns = sql_word[1] if sql_word[1] != '*' else "*"
                
                # 调用 join_tables 函数
                try:
                    result = dbms_function.join_tables(
                        tables=tables,
                        join_conditions=join_conditions,
                        using_db=using_db,
                        columns=columns,
                        where_condition=where_condition
                    )
                    
                    # 格式化输出结果
                    if result:
                        pt = PrettyTable()
                        pt.field_names = result[0]  # 第一行是表头
                        for row in result[1:]:  # 其余行是数据
                            pt.add_row(row)
                        return pt
                    else:
                        print("查询结果为空")
                        return None
                        
                except Exception as e:
                    print(f"查询执行出错: {str(e)}")
                    return None
            else:
                # 处理普通SELECT查询
                match = re.search(r'select\s+(.+?)\s+from\s+(\w+)(?:\s+where\s+(.+))?', sql.lower())
                if match:
                    columns = match.group(1)
                    table = match.group(2)
                    where_cond = match.group(3)
                    
                    if table not in using_db.sheetnames:
                        print(f"表 {table} 不存在")
                        return None
                        
                    sheet = using_db[table]
                    rows = list(dbms_function.iter_rows(sheet))
                    
                    # 创建结果表
                    pt = PrettyTable()
                    if columns == '*':
                        pt.field_names = rows[0]  # 使用第一行作为列名
                        result_rows = rows[1:]
                    else:
                        col_names = [c.strip() for c in columns.split(',')]
                        pt.field_names = col_names
                        col_indices = [rows[0].index(col) for col in col_names]
                        result_rows = [[row[i] for i in col_indices] for row in rows[1:]]
                    
                    # 处理WHERE条件
                    if where_cond:
                        filtered_rows = []
                        for row in result_rows:
                            if dbms_function.eval_where_condition(row, where_cond, rows[0]):
                                filtered_rows.append(row)
                        result_rows = filtered_rows
                    
                    for row in result_rows:
                        pt.add_row(row)
                        
                    print(pt)
                    return pt
                else:
                    print("[!] Invalid SELECT syntax")
                    return None
        except Exception as e:
            print(f"查询执行出错: {str(e)}")
            return None

def run():
    global user
    user = dbms_function.login(user)
    while True:
        command = get_command()
        if command == 'quit' or command == 'exit':
            print("[🍻] Thanks for using Mini DBMS. Bye~~")
            exit(0)
        elif command == 'help':
            help()
        else:
            query(command)

def userLogin(username, password, flagFirst, flagLogin):
    global user
    user, flagFirst, flagLogin = dbms_function.login(user, username, password, flagFirst, flagLogin)
    return flagFirst, flagLogin

def interpreter(command):
    print(f"正在执行命令: {command}")
    
    if command == 'quit' or command == 'exit':
        print("感谢使用迷你版DBMS！")
        exit(0)
    elif command == 'help':
        help()
    else:
        try:
            result = query(command)
            print("命令执行完成")
            return result
        except Exception as e:
            print(f"执行出错: {str(e)}")
            raise e
    f = os.open('data/log.txt', os.O_RDWR | os.O_APPEND)
    byte_str = ((time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())) + ' ' + command + '\n').encode('utf-8')
    os.write(f, byte_str)
