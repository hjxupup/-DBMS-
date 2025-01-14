import os
import sys
#导入QFileDialog
from PyQt5.QtWidgets import QFileDialog

import qdarkstyle
from PyQt5 import QtGui, QtWidgets
from PyQt5.QtCore import pyqtSignal, QObject
from PyQt5.QtGui import QTextCursor
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtWidgets import QTableWidgetItem
from openpyxl import load_workbook
from prettytable import PrettyTable

import UI_lan
import main_interpreter
from utils import assign_style_qt, get_merge_cell_list
from ai_correction import SQLCorrector

username = ''

password = ''
flagFirst = True
flagLogin = False
filePath = 'data/'

class Stream(QObject):
    """Redirects console output to text widget."""
    newText = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self._is_valid = True

    def write(self, text):
        if self._is_valid:
            try:
                QtWidgets.QApplication.processEvents()
                self.newText.emit(str(text))
            except RuntimeError:
                self._is_valid = False

    def __del__(self):
        self._is_valid = False

class anaxcelhandler(QtWidgets.QMainWindow, UI_lan.Ui_MainWindow):

    def __init__(self, parent=None):
        super(anaxcelhandler, self).__init__(parent)
        if getattr(sys, 'frozen', False):
            self.frozen = 'ever so'
            self.bundle_dir = sys._MEIPASS
        else:
            self.bundle_dir = os.path.dirname(os.path.abspath(__file__))
        self.setupUi(self)
        # self.setWindowIcon(QtGui.QIcon(self.bundle_dir + '/icons/icon.png'))
        # self.setStyleSheet(open("Dark/darkstyle.qss", "r").read())
        # self.setStyleSheet(open("qss/1.qss", "r").read())

        self.listWidget.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.pushButtonbrowse.clicked.connect(self.openFileNamesDialog)
        self.pushButtonbrowseLoad.clicked.connect(self.openFileNamesDialogLoad)
        self.pushButtonclear.clicked.connect(self.clearwidget)
        self.pushButtonselall.clicked.connect(self.selectall)
        self.pushButtonload.clicked.connect(self.LoadProcess)
        self.pushButton_submit.clicked.connect(self.submit)

        self.comboBoxfiletype.addItems(['xlsx', 'xls'])

        # ==========log=====
        self.stream = Stream()
        self.stream.newText.connect(self.onUpdateText)
        sys.stdout = self.stream

        # ==========show====
        self.flag_confirm = False
        self.activate_file = [None, None]
        self.comboBox_wb.activated.connect(self.wbActivated)
        self.comboBox_ws.activated.connect(self.wsActivated)
        self.tableWidget.itemClicked.connect(self.handleItemClick)
        #self.pushButton_clear_idx.clicked.connect(self.clear_idx)
        #self.pushButton_confirm_idx.clicked.connect(self.confirm_idx)
        # ==========show====

        # ==========context===
        self.infos = {}
        self.infos_bak = {}

        # 创建SQL纠错器实例
        self.sql_corrector = SQLCorrector()
        
        # 连接输入框的文本变化信号
        self.lineEdit_input.textChanged.connect(self.on_input_changed)

        # 添加回车事件处理方法
        self.lineEdit_input.returnPressed.connect(self.on_return_pressed)

    def submit(self):
        global username
        global password
        global flagFirst
        global flagLogin
        
        # 获取输入内容
        input_text = self.lineEdit_input.text().strip()
        
        if flagFirst:
            if not input_text:
                print('\n请输入用户名:')
                return
            print('请输入用户名:')
            flagFirst = False
        elif username == '':
            if not input_text:
                print('\n无法读取用户名，请重新输入:')
                return
            username = input_text
            print(username)
            print('密码:')
        elif not username == '' and password == '':
            if not input_text:
                print('\n无法读取密码，请重新输入:')
                return
            password = input_text
            print(password)
            print('再次点击进行验证')
        elif not username == '' and not password == '' and not flagLogin:
            flagFirst, flagLogin = main_interpreter.userLogin(username, password, flagFirst, flagLogin)
            if not flagLogin:
                username = ''
                password = ''
        elif flagLogin:
            if not input_text:
                # 在AI提示框中显示提示
                self.textBrowser_ai.clear()
                self.textBrowser_ai.append("<b>提示:</b>")
                self.textBrowser_ai.append("• 请输入需要执行的SQL操作")
                self.textBrowser_ai.append("• 输入help获取帮助信息")
                return
            
            sql = input_text
            # 在执行前进行语法检查
            corrected_sql, corrections = self.sql_corrector.correct_sql(sql)
            
            # 打印原始SQL和纠正后的SQL，用于调试
            print(f"原始SQL: {sql}")
            print(f"纠正后的SQL: {corrected_sql}")
            
            if corrections:
                # 显示纠正建议
                self.textBrowser_ai.clear()
                self.textBrowser_ai.append("<b>发现以下问题:</b>")
                for correction in corrections:
                    self.textBrowser_ai.append(f"• {correction}")
                    
                # 询问是否使用纠正后的SQL
                reply = QtWidgets.QMessageBox.question(
                    self, 
                    'SQL纠正',
                    f'是否使用纠正后的SQL:\n{corrected_sql}',
                    QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No
                )
                
                if reply == QtWidgets.QMessageBox.Yes:
                    sql = corrected_sql
            
            try:
                # 执行SQL并捕获可能的异常
                result = main_interpreter.interpreter(sql)
                
                # 显示结果
                if result:
                    if isinstance(result, PrettyTable):
                        # 格式化显示
                        formatted_result = str(result)
                        self.textBrowserlog.clear()
                        self.textBrowserlog.append(formatted_result)
                    else:
                        self.textBrowserlog.clear()
                        self.textBrowserlog.append(str(result))
                
            except Exception as e:
                # 显示错误信息
                self.textBrowser_ai.clear()
                self.textBrowser_ai.append("<b>执行出错:</b>")
                self.textBrowser_ai.append(f"• {str(e)}")
                print(f"执行错误: {str(e)}")

    def use_palette(self):
        self.setWindowTitle("设置背景图片")
        window_pale = QtGui.QPalette()
        window_pale.setBrush(self.backgroundRole(), QtGui.QBrush(QtGui.QPixmap(self.bundle_dir + '/icons/bg.jpeg')))
        self.setPalette(window_pale)

    def onUpdateText(self, text):
        """Write console output to text widget."""
        cursor = self.textBrowserlog.textCursor()
        cursor.movePosition(QTextCursor.End)
        cursor.insertText(text)
        self.textBrowserlog.setTextCursor(cursor)
        self.textBrowserlog.ensureCursorVisible()

    def openFileNamesDialog(self):
        # 获取filePath所有xlsx文件存储到files中
        global filePath
        files = []
        for file in os.listdir(filePath):
            if file.endswith('.xlsx'):
                files.append(file)

        """options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        filterxls = "XLS (*.xls *.XLS)"
        filterxlsx = "XLSX (*.xlsx *.XLSX)"
        print('打开文件')
        if self.comboBoxfiletype.currentIndex() == 1:
            files, _ = QFileDialog.getOpenFileNames(self, "Select XLS Files", filter=filterxls, options=options)
            if files:
                for file in files:
                    #仅保留file的文件名且去除后缀
                    self.listWidget.addItem(os.path.basename(file).split('.')[0])
        elif self.comboBoxfiletype.currentIndex() == 0:
            files, _ = QFileDialog.getOpenFileNames(self, "Select XLSX Files", filter=filterxlsx, options=options)
            if files:
                for file in files:
                    #仅保留file的文件名且去除后缀
                    self.listWidget.addItem(os.path.basename(file).split('.')[0])
        """
        if files:
            for file in files:
                # 仅保留file的文件名且去除后缀
                self.listWidget.addItem(os.path.basename(file).split('.')[0])

    def openFileNamesDialogLoad(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        filterxls = "XLS (*.xls *.XLS)"
        filterxlsx = "XLSX (*.xlsx *.XLSX)"
        print('打开文件')
        if self.comboBoxfiletype.currentIndex() == 1:
            files, _ = QFileDialog.getOpenFileNames(self, "Select XLS Files", filter=filterxls, options=options)
            if files:
                for file in files:
                    self.listWidget.addItem(os.path.basename(file).split('.')[0])
        elif self.comboBoxfiletype.currentIndex() == 0:
            files, _ = QFileDialog.getOpenFileNames(self, "Select XLSX Files", filter=filterxlsx, options=options)
            if files:
                for file in files:
                    self.listWidget.addItem(os.path.basename(file).split('.')[0])

    def clearwidget(self):
        self.listWidget.clear()
        self.tableWidget.clear()
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
        self.comboBox_x.clear()
        self.comboBox_y.clear()
        self.comboBox_wb.clear()
        self.comboBox_ws.clear()
        self.comboBox_r1.clear()
        self.comboBox_r2.clear()

    def clearcontext_all(self):
        self.tableWidget.clear()
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
        self.comboBox_x.clear()
        self.comboBox_y.clear()
        self.comboBox_wb.clear()
        self.comboBox_ws.clear()
        self.comboBox_r1.clear()
        self.comboBox_r2.clear()

    def clearcontext_show(self):
        self.tableWidget.clear()
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)

    def clear_idx(self):
        self.comboBox_x.clear()
        self.comboBox_y.clear()
        self.comboBox_r1.clear()
        self.comboBox_r2.clear()

    def assign_dict(self, dict1, dict2):
        for k, v in dict1.items():
            if isinstance(v, dict):
                dict_tmp = dict()
                dict2[k] = self.assign_dict(v, dict_tmp)
            else:
                dict2[k] = v
        return dict2

    def confirm_idx(self):
        self.infos_bak = self.assign_dict(self.infos, self.infos_bak)

        x = self.comboBox_x.itemText(self.comboBox_x.currentIndex())
        y = self.comboBox_y.itemText(self.comboBox_y.currentIndex())

        r1 = self.comboBox_r1.itemText(self.comboBox_r1.currentIndex())
        r2 = self.comboBox_r2.itemText(self.comboBox_r2.currentIndex())

        wb = self.comboBox_wb.itemText(self.comboBox_wb.currentIndex())
        ws = self.comboBox_ws.itemText(self.comboBox_ws.currentIndex())

        if wb == '' or ws == '':
            QMessageBox.about(self, "hi,Mini DBMS by group 10", '先load文件')
        else:
            x = int(x) if x != '' else x
            y = int(y) if y != '' else y
            r1 = int(r1) if r1 != '' else r1
            r2 = int(r2) if r2 != '' else r2

            if self.checkBox_book.isChecked():
                print('book')
                key_idx = [x, y]
                rg = [r1, 'last']
                for wb_k in self.infos_bak.keys():
                    ws_keys = self.infos_bak[wb_k]['sheet_names']
                    for ws_k in ws_keys.keys():
                        self.infos_bak[wb_k]['sheet_names'][ws_k] = [key_idx, rg]
            elif self.checkBox_sheet.isChecked():
                print('sheet')
                key_idx = [x, y]
                rg = [r1, 'last']
                ws_keys = self.infos_bak[wb]['sheet_names']
                for ws_k in ws_keys.keys():
                    self.infos_bak[wb]['sheet_names'][ws_k] = [key_idx, rg]
            else:
                print('cell')
                key_idx = [x, y]
                rg = [r1, r2]
                self.infos_bak[wb]['sheet_names'][ws] = [key_idx, rg]
            self.flag_confirm = True

    def selectall(self):
        self.listWidget.selectAll()
        items = self.listWidget.selectedItems()
        if len(items) == 0:
            QMessageBox.about(self, "hi,Mini DBMS by group 10", '请先加载文件')

    def LoadProcess(self):
        self.clearcontext_all()
        if self.comboBoxfiletype.currentIndex() == 1:  # xls
            QMessageBox.about(self, "hi,Mini DBMS by group 10", '不支持 xls 格式文件')
        elif self.comboBoxfiletype.currentIndex() == 0:  # xlsx
            items = self.listWidget.selectedItems()
            if len(items) == 0:
                QMessageBox.about(self, "hi,Mini DBMS by group 10", '请先选择文件')
            else:
                self.infos = {}
                for i in list(items):
                    file_path = str(filePath + i.text() + '.xlsx')
                    wb = load_workbook(filename=file_path)
                    name = os.path.split(file_path)[-1]

                    sheet_names = wb.sheetnames

                    sheets_dict = {}
                    for s in sheet_names:
                        sheets_dict[s] = []
                    self.infos[name] = {'path': file_path, 'sheet_names': sheets_dict}
                    wb.close()
                for k in self.infos.keys():
                    self.comboBox_wb.addItem(k)
                k = self.comboBox_wb.itemText(0)
                sheets = list(self.infos[k]['sheet_names'].keys())
                for s in sheets:
                    self.comboBox_ws.addItem(s)
                self.activate_file[0] = self.infos[k]['path']
                self.activate_file[1] = list(self.infos[k]['sheet_names'].keys())[0]

                self.show_excel()
        print('可以预览文件')

    def wbActivated(self, index):
        self.clearcontext_show()
        wb_k = self.comboBox_wb.itemText(index)
        sheets = list(self.infos[wb_k]['sheet_names'].keys())
        self.comboBox_ws.clear()
        for s in sheets:
            self.comboBox_ws.addItem(s)
        self.activate_file[0] = self.infos[wb_k]['path']
        self.activate_file[1] = list(self.infos[wb_k]['sheet_names'].keys())[0]
        self.show_excel()

    def wsActivated(self, index):
        ws_k = self.comboBox_ws.itemText(index)
        self.activate_file[1] = ws_k
        self.show_excel()

    def handleItemClick(self, item):
        cont = item.text()
        self.comboBox_x.clear()
        self.comboBox_y.clear()
        self.comboBox_r1.clear()
        row = item.row() + 1
        column = item.column() + 1
        # =======对合并的单元格取idx
        for p in self.merge_position:
            if row == p[0] and column == p[1]:
                row = row + (p[2] - p[0])
                break
        # =======对合并的单元格取idx
        self.comboBox_x.addItem(str(row))
        self.comboBox_y.addItem(str(column))
        self.comboBox_r1.addItem(str(row + 1))

    def show_excel(self):
        self.merge_position = []
        path = self.activate_file[0]
        sheetname = self.activate_file[1]
        wb = load_workbook(filename=path)
        ws = wb[sheetname]
        num_row = ws.max_row
        num_column = ws.max_column
        self.tableWidget.setColumnCount(num_column)
        self.tableWidget.setRowCount(num_row)

        # ======合并单元格=======
        merge_idx = ws.merged_cells
        merge_idx = get_merge_cell_list(merge_idx)

        for i in range(len(merge_idx)):
            m_idx = merge_idx[i]
            self.tableWidget.setSpan(m_idx[0] - 1, m_idx[1] - 1, m_idx[2] - m_idx[0] + 1, m_idx[3] - m_idx[1] + 1)
            self.merge_position.append([m_idx[0], m_idx[1], m_idx[2]])  # [x1,y1,range]
        # ======合并单元格=======

        # ======单元格大小=======
        for i in range(1, num_row + 1):
            h = ws.row_dimensions[i].height
            if h is not None:
                self.tableWidget.setRowHeight(i - 1, h)
        # for i in range(1,num_column+1):
        #     w = ws.column_dimensions[get_column_letter(i)].width
        #     if w is not None:
        #         self.tableWidget.setColumnWidth(i-1,w)
        # ======单元格大小=======

        self.comboBox_r2.clear()
        for i in range(1, num_row + 1):
            self.comboBox_r2.addItem(str(num_row - i + 1))
            row_sizes = []
            for j in range(1, num_column + 1):
                cell = ws.cell(row=i, column=j)
                if cell.value is not None:
                    item = QTableWidgetItem(str(cell.value))
                    assign_style_qt(item, cell)
                else:
                    item = QTableWidgetItem()
                self.tableWidget.setItem(i - 1, j - 1, item)

        # self.tableWidget.resizeColumnsToContents()
        # self.tableWidget.resizeRowsToContents()
        wb.close()

    def on_input_changed(self):
        """当输入框文本变化时更新AI提示"""
        sql = self.lineEdit_input.text().strip()
        
        # 清空之前的提示
        self.textBrowser_ai.clear()
        
        if not sql:
            # 输入为空时显示基本提示
            self.textBrowser_ai.append("<b>基本提示:</b>")
            self.textBrowser_ai.append("• 请输入SQL命令")
            self.textBrowser_ai.append("• 输入help获取帮助信息")
            self.textBrowser_ai.append("")
            self.textBrowser_ai.append("<b>常用命令:</b>")
            self.textBrowser_ai.append("• create database {name}")
            self.textBrowser_ai.append("• use database {name}")
            self.textBrowser_ai.append("• create table {name} (...)")
            self.textBrowser_ai.append("• select * from {table}")
            return
            
        # 获取AI建议
        corrected_sql, corrections = self.sql_corrector.correct_sql(sql)
        
        # 如果有语法纠正
        if corrected_sql != sql:
            self.textBrowser_ai.append("<b>语法纠正建议:</b>")
            self.textBrowser_ai.append(f"• 将 '{sql}' 修改为 '{corrected_sql}'")
            self.textBrowser_ai.append("")
            
        # 添加其他AI建议
        if 'join' in sql.lower():
            # 获取JOIN相关建议
            join_suggestions = self.sql_corrector.suggest_join_optimizations(sql)
            if join_suggestions:
                self.textBrowser_ai.append("<b>JOIN查询优化建议:</b>")
                for suggestion in join_suggestions:
                    self.textBrowser_ai.append(f"• {suggestion}")
                self.textBrowser_ai.append("")
                
        # 获取索引建议
        index_suggestions = self.sql_corrector.suggest_indexes(sql)
        if index_suggestions:
            self.textBrowser_ai.append("<b>索引优化建议:</b>")
            for suggestion in index_suggestions:
                self.textBrowser_ai.append(f"• {suggestion}")
            self.textBrowser_ai.append("")
            
        # 获取AI优化建议
        ai_suggestions = self.sql_corrector.suggest_ai_optimizations(sql)
        if ai_suggestions:
            self.textBrowser_ai.append("<b>智能优化建议:</b>")
            for suggestion in ai_suggestions:
                self.textBrowser_ai.append(f"• {suggestion}")
                

    def on_return_pressed(self):
        """处理回车键事件"""
        self.submit()  # 直接调用submit方法
        # 清空输入框
        self.lineEdit_input.clear()

    def closeEvent(self, event):
        """处理窗口关闭事件"""
        sys.stdout = sys.__stdout__  # 恢复标准输出
        self.stream._is_valid = False
        super().closeEvent(event)

if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    window = anaxcelhandler()
    window.openFileNamesDialog()
    app.setStyleSheet(qdarkstyle.load_stylesheet_pyqt5())
    window.show()
    try:
        sys.exit(app.exec_())
    finally:
        # 确保恢复标准输出
        sys.stdout = sys.__stdout__
